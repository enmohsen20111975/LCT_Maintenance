from typing import List, Dict, Any, Optional, Tuple
from sqlalchemy import Table, text, inspect, desc, asc
from sqlalchemy.exc import SQLAlchemyError
from models import db
import pandas as pd
import os
import tempfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import numpy as np
import json

class RelationshipService:
    """Service for managing table relationships and joins."""
    
    def __init__(self):
        """Initialize the relationship service."""
        pass
    
    def _convert_numpy_types(self, obj):
        """Convert numpy types to Python native types for JSON serialization."""
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, dict):
            return {key: self._convert_numpy_types(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._convert_numpy_types(item) for item in obj]
        else:
            return obj
    
    def get_table_columns(self, table_name: str, database: str = 'excel_data') -> List[Dict[str, Any]]:
        """Get columns for a specific table with their data types."""
        try:
            if database and database != 'excel_data':
                # Connect to specific database
                import sqlite3
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', f'{database}.db')
                
                if not os.path.exists(db_path):
                    return []
                
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                
                # Check if table exists
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                if not cursor.fetchone():
                    conn.close()
                    return []
                
                # Get column info from sqlite
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()
                
                columns = [
                    {
                        'name': col[1],
                        'type': col[2],
                        'nullable': not col[3],
                        'primary_key': col[5] == 1
                    }
                    for col in columns_info
                ]
                
                conn.close()
                return columns
            else:
                # Use default database connection
                inspector = inspect(db.engine)
                if table_name not in inspector.get_table_names():
                    return []
                
                columns = inspector.get_columns(table_name)
                return [
                    {
                        'name': col['name'],
                        'type': str(col['type']),
                        'nullable': col.get('nullable', True),
                        'primary_key': col.get('primary_key', False)
                    }
                    for col in columns
                ]
        except Exception as e:
            print(f"Error getting columns for {table_name}: {e}")
            return []
    
    def find_potential_relationships(self, tables: List[str]) -> List[Dict[str, Any]]:
        """Find potential relationships between tables based on column names and data types."""
        relationships = []
        
        if len(tables) < 2:
            return relationships
        
        # Get column information for all tables
        table_columns = {}
        for table in tables:
            table_columns[table] = self.get_table_columns(table)
        
        # Compare each pair of tables
        for i, table1 in enumerate(tables):
            for table2 in tables[i+1:]:
                cols1 = table_columns[table1]
                cols2 = table_columns[table2]
                
                # Find matching column names and compatible types
                for col1 in cols1:
                    for col2 in cols2:
                        if self._columns_compatible(col1, col2):
                            confidence = self._calculate_confidence(col1, col2)
                            
                            relationships.append({
                                'table1': table1,
                                'column1': col1['name'],
                                'table2': table2,
                                'column2': col2['name'],
                                'type1': col1['type'],
                                'type2': col2['type'],
                                'confidence': confidence,
                                'suggested_join': 'INNER' if confidence > 0.8 else 'LEFT'
                            })
        
        # Sort by confidence (highest first)
        relationships.sort(key=lambda x: x['confidence'], reverse=True)
        return relationships
    
    def _columns_compatible(self, col1: Dict[str, Any], col2: Dict[str, Any]) -> bool:
        """Check if two columns are compatible for joining."""
        # Exact name match
        if col1['name'].lower() == col2['name'].lower():
            return True
        
        # Common ID patterns
        id_patterns = [
            (col1['name'].lower().endswith('_id') and col2['name'].lower() == 'id'),
            (col2['name'].lower().endswith('_id') and col1['name'].lower() == 'id'),
            (col1['name'].lower() == 'id' and col2['name'].lower() == 'id'),
        ]
        
        if any(id_patterns):
            return True
        
        # Similar names (fuzzy matching)
        name1_clean = col1['name'].lower().replace('_', '').replace(' ', '')
        name2_clean = col2['name'].lower().replace('_', '').replace(' ', '')
        
        if name1_clean == name2_clean:
            return True
        
        return False
    
    def _calculate_confidence(self, col1: Dict[str, Any], col2: Dict[str, Any]) -> float:
        """Calculate confidence score for a potential relationship."""
        confidence = 0.0
        
        # Exact name match gets highest score
        if col1['name'].lower() == col2['name'].lower():
            confidence += 0.9
        
        # ID patterns get high scores
        if col1['name'].lower() == 'id' or col2['name'].lower() == 'id':
            confidence += 0.8
        elif '_id' in col1['name'].lower() or '_id' in col2['name'].lower():
            confidence += 0.7
        
        # Similar names
        name1_clean = col1['name'].lower().replace('_', '').replace(' ', '')
        name2_clean = col2['name'].lower().replace('_', '').replace(' ', '')
        if name1_clean == name2_clean:
            confidence += 0.6
        
        # Both primary keys is less likely to be a good relationship
        if col1.get('primary_key') and col2.get('primary_key'):
            confidence -= 0.3
        
        return min(confidence, 1.0)
    
    def validate_join_configuration(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Validate a join configuration and return validation results."""
        errors = []
        warnings = []
        
        tables = config.get('tables', [])
        joins = config.get('joins', [])
        
        # Check basic requirements
        if not tables:
            errors.append("No tables specified")
            return {'valid': False, 'errors': errors, 'warnings': warnings}
        
        # Validate tables exist
        inspector = inspect(db.engine)
        existing_tables = inspector.get_table_names()
        
        for table in tables:
            if table not in existing_tables:
                errors.append(f"Table '{table}' does not exist")
        
        # For single table, joins are not required
        if len(tables) == 1:
            return {'valid': True, 'errors': errors, 'warnings': warnings}
        
        # For multiple tables, validate joins
        if len(tables) > 1 and not joins:
            errors.append("Multiple tables selected but no joins defined")
        
        # Validate joins
        for i, join in enumerate(joins):
            table1 = join.get('table1')
            table2 = join.get('table2')
            column1 = join.get('column1')
            column2 = join.get('column2')
            join_type = join.get('join_type', 'INNER')
            
            # Check required fields
            if not all([table1, table2, column1, column2]):
                errors.append(f"Join {i+1}: Missing required fields")
                continue
            
            # Check tables are in the selected list
            if table1 not in tables:
                errors.append(f"Join {i+1}: Table '{table1}' not in selected tables")
            if table2 not in tables:
                errors.append(f"Join {i+1}: Table '{table2}' not in selected tables")
            
            # Check columns exist
            if table1 in existing_tables:
                cols1 = [col['name'] for col in self.get_table_columns(table1)]
                if column1 not in cols1:
                    errors.append(f"Join {i+1}: Column '{column1}' not found in table '{table1}'")
            
            if table2 in existing_tables:
                cols2 = [col['name'] for col in self.get_table_columns(table2)]
                if column2 not in cols2:
                    errors.append(f"Join {i+1}: Column '{column2}' not found in table '{table2}'")
            
            # Validate join type
            valid_joins = ['INNER', 'LEFT', 'RIGHT', 'FULL OUTER']
            if join_type.upper() not in valid_joins:
                errors.append(f"Join {i+1}: Invalid join type '{join_type}'")
        
        # Check for circular dependencies or disconnected tables
        if len(tables) > 1 and len(joins) < len(tables) - 1:
            warnings.append("Some tables may not be connected through joins")
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }
    
    def preview_joined_data(self, config: Dict[str, Any], limit: int = 100) -> Dict[str, Any]:
        """Generate a comprehensive preview of the joined data based on configuration."""
        try:
            validation = self.validate_join_configuration(config)
            if not validation['valid']:
                return {
                    'success': False,
                    'error': 'Configuration validation failed',
                    'details': validation['errors']
                }
            
            query = self._build_join_query(config, limit=limit)
            
            result = db.session.execute(text(query))
            columns = result.keys()
            rows = result.fetchall()
            
            # Convert to list of dictionaries
            data = []
            for row in rows:
                row_dict = {}
                for i, value in enumerate(row):
                    row_dict[columns[i]] = value
                data.append(row_dict)
            
            # Get comprehensive preview analysis
            preview_analysis = self._generate_comprehensive_preview_analysis(config, data, query)
            
            return {
                'success': True,
                'data': data,
                'columns': list(columns),
                'row_count': len(data),
                'query': query,
                'statistics': preview_analysis['statistics'],
                'table_info': preview_analysis['table_info'],
                'join_analysis': preview_analysis['join_analysis'],
                'output_analysis': preview_analysis['output_analysis'],
                'data_quality': preview_analysis['data_quality'],
                'preview_insights': preview_analysis['preview_insights'],
                'export_recommendations': preview_analysis['export_recommendations']
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _generate_comprehensive_preview_analysis(self, config: Dict[str, Any], preview_data: List[Dict], query: str) -> Dict[str, Any]:
        """Generate comprehensive analysis of the preview data for enhanced output visualization."""
        try:
            analysis = {
                'statistics': self._get_enhanced_preview_statistics(config, preview_data),
                'table_info': self._get_tables_info(config.get('tables', [])),
                'join_analysis': self._analyze_joins(config),
                'output_analysis': self._analyze_output_structure(preview_data, config),
                'data_quality': self._assess_data_quality(preview_data),
                'preview_insights': self._generate_preview_insights(preview_data, config),
                'export_recommendations': self._generate_export_recommendations(preview_data, config)
            }
            
            return analysis
            
        except Exception as e:
            return {
                'statistics': {'error': str(e)},
                'table_info': {},
                'join_analysis': {},
                'output_analysis': {},
                'data_quality': {},
                'preview_insights': [],
                'export_recommendations': []
            }
    
    def _get_enhanced_preview_statistics(self, config: Dict[str, Any], preview_data: List[Dict]) -> Dict[str, Any]:
        """Get enhanced statistics about the preview data."""
        try:
            if not preview_data:
                return {'preview_row_count': 0, 'columns': 0}
            
            stats = {
                'preview_row_count': len(preview_data),
                'total_columns': len(preview_data[0].keys()) if preview_data else 0,
                'column_analysis': {},
                'data_distribution': {},
                'memory_estimation': {},
                'performance_metrics': {}
            }
            
            # Enhanced column analysis
            for column in preview_data[0].keys():
                values = [row.get(column) for row in preview_data]
                non_null_values = [v for v in values if v is not None and v != '']
                
                # Basic statistics
                column_stats = {
                    'total_values': len(values),
                    'non_null_count': len(non_null_values),
                    'null_count': len(values) - len(non_null_values),
                    'null_percentage': round((len(values) - len(non_null_values)) / len(values) * 100, 2) if values else 0,
                    'unique_count': len(set(str(v) for v in non_null_values)) if non_null_values else 0
                }
                
                # Data type inference and analysis
                if non_null_values:
                    column_stats.update(self._analyze_column_data_type(non_null_values))
                    column_stats['sample_values'] = list(set(str(v) for v in non_null_values))[:5]
                    column_stats['uniqueness_ratio'] = round(column_stats['unique_count'] / len(non_null_values), 3) if non_null_values else 0
                else:
                    column_stats.update({
                        'inferred_type': 'unknown',
                        'sample_values': [],
                        'uniqueness_ratio': 0
                    })
                
                stats['column_analysis'][column] = column_stats
            
            # Get total row counts for each table
            total_counts = {}
            estimated_full_result_size = 0
            
            for table in config.get('tables', []):
                try:
                    count_result = db.session.execute(text(f"SELECT COUNT(*) FROM {table}"))
                    total_counts[table] = count_result.scalar()
                    if len(config.get('tables', [])) == 1:
                        estimated_full_result_size = total_counts[table]
                except Exception as e:
                    total_counts[table] = 0
            
            # Estimate full result size for joins
            if len(config.get('tables', [])) > 1:
                estimated_full_result_size = self._estimate_join_result_size(config, total_counts)
            
            stats['total_row_counts'] = total_counts
            stats['estimated_full_result_size'] = estimated_full_result_size
            
            # Memory estimation
            stats['memory_estimation'] = self._estimate_memory_usage(preview_data, estimated_full_result_size)
            
            return stats
            
        except Exception as e:
            return {'error': str(e), 'preview_row_count': 0}
    
    def _analyze_column_data_type(self, values: List) -> Dict[str, Any]:
        """Analyze column values to infer data type and characteristics."""
        if not values:
            return {'inferred_type': 'unknown'}
        
        # Convert to strings for analysis
        str_values = [str(v).strip() for v in values if v is not None]
        
        # Check for numeric types
        numeric_count = 0
        date_count = 0
        boolean_count = 0
        
        for val in str_values[:min(100, len(str_values))]:  # Sample for performance
            # Check if numeric
            try:
                float(val)
                numeric_count += 1
                continue
            except:
                pass
            
            # Check if date
            if self._is_date_like(val):
                date_count += 1
                continue
            
            # Check if boolean
            if val.lower() in ['true', 'false', 'yes', 'no', '1', '0', 't', 'f', 'y', 'n']:
                boolean_count += 1
        
        sample_size = min(100, len(str_values))
        
        # Determine primary type
        if numeric_count / sample_size > 0.8:
            return {
                'inferred_type': 'numeric',
                'numeric_ratio': round(numeric_count / sample_size, 3),
                'min_length': min(len(str(v)) for v in values),
                'max_length': max(len(str(v)) for v in values),
                'avg_length': round(sum(len(str(v)) for v in values) / len(values), 1)
            }
        elif date_count / sample_size > 0.6:
            return {
                'inferred_type': 'date',
                'date_ratio': round(date_count / sample_size, 3),
                'min_length': min(len(str(v)) for v in values),
                'max_length': max(len(str(v)) for v in values)
            }
        elif boolean_count / sample_size > 0.7:
            return {
                'inferred_type': 'boolean',
                'boolean_ratio': round(boolean_count / sample_size, 3)
            }
        else:
            return {
                'inferred_type': 'text',
                'min_length': min(len(str(v)) for v in values),
                'max_length': max(len(str(v)) for v in values),
                'avg_length': round(sum(len(str(v)) for v in values) / len(values), 1)
            }
    
    def _is_date_like(self, value: str) -> bool:
        """Check if a string value looks like a date."""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
            r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
            r'\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
            r'\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD
        ]
        
        import re
        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return True
        return False
    
    def _estimate_join_result_size(self, config: Dict[str, Any], table_counts: Dict[str, int]) -> int:
        """Estimate the size of the full join result."""
        tables = config.get('tables', [])
        joins = config.get('joins', [])
        
        if len(tables) <= 1:
            return table_counts.get(tables[0], 0) if tables else 0
        
        # Simple heuristic: start with smallest table and multiply by estimated join factors
        base_size = min(table_counts.values()) if table_counts.values() else 1000
        
        # Each join potentially multiplies the result
        for join in joins:
            join_type = join.get('join_type', 'INNER')
            if join_type.upper() == 'INNER':
                # Inner join typically reduces size
                base_size = int(base_size * 0.7)
            elif join_type.upper() in ['LEFT', 'RIGHT']:
                # Outer joins may increase size
                base_size = int(base_size * 1.2)
            else:
                # Full outer join could significantly increase
                base_size = int(base_size * 1.5)
        
        return max(base_size, 0)
    
    def _estimate_memory_usage(self, preview_data: List[Dict], estimated_full_size: int) -> Dict[str, Any]:
        """Estimate memory usage for the full dataset."""
        if not preview_data:
            return {'estimated_mb': 0, 'preview_kb': 0}
        
        # Calculate average row size in bytes
        sample_row = preview_data[0] if preview_data else {}
        avg_row_size = 0
        
        for value in sample_row.values():
            if value is not None:
                avg_row_size += len(str(value)) * 2  # Rough estimate including overhead
        
        preview_size_kb = (len(preview_data) * avg_row_size) / 1024
        estimated_full_mb = (estimated_full_size * avg_row_size) / (1024 * 1024)
        
        return {
            'preview_kb': round(preview_size_kb, 2),
            'estimated_mb': round(estimated_full_mb, 2),
            'avg_row_size_bytes': avg_row_size,
            'performance_warning': estimated_full_mb > 100
        }
    
    def _analyze_output_structure(self, preview_data: List[Dict], config: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze the structure and characteristics of the output data."""
        if not preview_data:
            return {'empty_result': True}
        
        analysis = {
            'structure_type': 'tabular',
            'column_groups': {},
            'relationships_preserved': [],
            'data_normalization': {},
            'output_characteristics': {}
        }
        
        # Group columns by source table
        tables = config.get('tables', [])
        for table in tables:
            table_columns = [col for col in preview_data[0].keys() if col.startswith(f'{table}_') or len(tables) == 1]
            if table_columns:
                analysis['column_groups'][table] = {
                    'columns': table_columns,
                    'column_count': len(table_columns)
                }
        
        # Analyze data normalization level
        total_columns = len(preview_data[0].keys())
        unique_ratios = []
        
        for column, values in [(k, [row[k] for row in preview_data]) for k in preview_data[0].keys()]:
            non_null = [v for v in values if v is not None]
            if non_null:
                unique_ratio = len(set(str(v) for v in non_null)) / len(non_null)
                unique_ratios.append(unique_ratio)
        
        avg_uniqueness = sum(unique_ratios) / len(unique_ratios) if unique_ratios else 0
        
        analysis['data_normalization'] = {
            'average_uniqueness': round(avg_uniqueness, 3),
            'normalization_level': 'high' if avg_uniqueness > 0.8 else 'medium' if avg_uniqueness > 0.3 else 'low',
            'potential_duplicates': avg_uniqueness < 0.5
        }
        
        # Output characteristics
        analysis['output_characteristics'] = {
            'total_columns': total_columns,
            'data_density': self._calculate_data_density(preview_data),
            'column_correlation': self._analyze_column_correlations(preview_data),
            'export_suitability': self._assess_export_suitability(preview_data)
        }
        
        return analysis
    
    def _calculate_data_density(self, preview_data: List[Dict]) -> float:
        """Calculate the density of non-null data."""
        if not preview_data:
            return 0.0
        
        total_cells = len(preview_data) * len(preview_data[0])
        filled_cells = 0
        
        for row in preview_data:
            for value in row.values():
                if value is not None and value != '':
                    filled_cells += 1
        
        return round(filled_cells / total_cells, 3) if total_cells > 0 else 0.0
    
    def _analyze_column_correlations(self, preview_data: List[Dict]) -> Dict[str, Any]:
        """Analyze potential correlations between columns."""
        if not preview_data or len(preview_data[0]) < 2:
            return {'correlation_count': 0, 'correlations': []}
        
        columns = list(preview_data[0].keys())
        correlations = []
        
        # Look for potential ID relationships
        for i, col1 in enumerate(columns):
            for col2 in columns[i+1:]:
                # Check if columns might be related (simple heuristic)
                values1 = [str(row[col1]) for row in preview_data if row[col1] is not None]
                values2 = [str(row[col2]) for row in preview_data if row[col2] is not None]
                
                if values1 and values2:
                    # Check for potential relationships
                    overlap = len(set(values1) & set(values2))
                    if overlap > 0:
                        correlation_strength = overlap / min(len(set(values1)), len(set(values2)))
                        if correlation_strength > 0.1:  # Threshold for meaningful correlation
                            correlations.append({
                                'column1': col1,
                                'column2': col2,
                                'strength': round(correlation_strength, 3),
                                'type': 'value_overlap'
                            })
        
        return {
            'correlation_count': len(correlations),
            'correlations': correlations[:5]  # Limit to top 5
        }
    
    def _assess_export_suitability(self, preview_data: List[Dict]) -> Dict[str, Any]:
        """Assess how suitable the data is for different export formats."""
        if not preview_data:
            return {'formats': {}}
        
        suitability = {
            'excel': {'score': 8, 'notes': []},
            'csv': {'score': 9, 'notes': []},
            'json': {'score': 7, 'notes': []},
            'pdf': {'score': 5, 'notes': []}
        }
        
        # Analyze characteristics that affect export suitability
        total_columns = len(preview_data[0])
        max_text_length = 0
        has_special_chars = False
        
        for row in preview_data:
            for value in row.values():
                if value is not None:
                    str_val = str(value)
                    max_text_length = max(max_text_length, len(str_val))
                    if any(char in str_val for char in ['\n', '\r', '\t', '"', "'"]):
                        has_special_chars = True
        
        # Adjust scores based on characteristics
        if total_columns > 20:
            suitability['excel']['score'] -= 2
            suitability['excel']['notes'].append('Many columns may require wide layout')
            suitability['pdf']['score'] -= 3
            suitability['pdf']['notes'].append('Too many columns for standard PDF layout')
        
        if max_text_length > 100:
            suitability['csv']['score'] -= 1
            suitability['csv']['notes'].append('Long text values may need special handling')
            suitability['excel']['score'] -= 1
            suitability['excel']['notes'].append('Long text may require cell wrapping')
        
        if has_special_chars:
            suitability['csv']['score'] -= 2
            suitability['csv']['notes'].append('Special characters require proper escaping')
        
        return {'formats': suitability}
    
    def _assess_data_quality(self, preview_data: List[Dict]) -> Dict[str, Any]:
        """Assess the quality of the preview data."""
        if not preview_data:
            return {'overall_score': 0, 'issues': [], 'recommendations': []}
        
        quality_assessment = {
            'overall_score': 10,
            'issues': [],
            'recommendations': [],
            'completeness': {},
            'consistency': {},
            'accuracy_indicators': {}
        }
        
        # Assess completeness
        total_cells = len(preview_data) * len(preview_data[0])
        null_cells = 0
        
        for row in preview_data:
            for value in row.values():
                if value is None or value == '':
                    null_cells += 1
        
        completeness_score = (total_cells - null_cells) / total_cells if total_cells > 0 else 0
        quality_assessment['completeness'] = {
            'score': round(completeness_score, 3),
            'null_percentage': round(null_cells / total_cells * 100, 2) if total_cells > 0 else 0
        }
        
        if completeness_score < 0.8:
            quality_assessment['overall_score'] -= 2
            quality_assessment['issues'].append('High percentage of missing values')
            quality_assessment['recommendations'].append('Consider data cleaning or imputation strategies')
        
        # Assess consistency
        inconsistencies = 0
        for column in preview_data[0].keys():
            values = [str(row[column]).strip() for row in preview_data if row[column] is not None]
            if values:
                # Check for inconsistent formatting
                if self._has_inconsistent_formatting(values):
                    inconsistencies += 1
        
        consistency_score = 1 - (inconsistencies / len(preview_data[0])) if preview_data[0] else 1
        quality_assessment['consistency'] = {
            'score': round(consistency_score, 3),
            'inconsistent_columns': inconsistencies
        }
        
        if consistency_score < 0.7:
            quality_assessment['overall_score'] -= 1
            quality_assessment['issues'].append('Inconsistent data formatting detected')
            quality_assessment['recommendations'].append('Standardize data formats before export')
        
        # Look for potential accuracy issues
        accuracy_issues = self._detect_accuracy_issues(preview_data)
        quality_assessment['accuracy_indicators'] = accuracy_issues
        
        if accuracy_issues['suspicious_patterns'] > 0:
            quality_assessment['overall_score'] -= 1
            quality_assessment['issues'].append('Suspicious data patterns detected')
        
        quality_assessment['overall_score'] = max(0, quality_assessment['overall_score'])
        
        return quality_assessment
    
    def _has_inconsistent_formatting(self, values: List[str]) -> bool:
        """Check if values have inconsistent formatting."""
        if len(values) < 2:
            return False
        
        # Check date formatting inconsistencies
        date_formats = set()
        for value in values[:20]:  # Sample for performance
            if self._is_date_like(value):
                if '-' in value:
                    date_formats.add('dash')
                elif '/' in value:
                    date_formats.add('slash')
        
        # Check case inconsistencies in text
        case_patterns = set()
        for value in values[:20]:
            if value.isalpha():
                if value.isupper():
                    case_patterns.add('upper')
                elif value.islower():
                    case_patterns.add('lower')
                elif value.istitle():
                    case_patterns.add('title')
                else:
                    case_patterns.add('mixed')
        
        return len(date_formats) > 1 or len(case_patterns) > 2
    
    def _detect_accuracy_issues(self, preview_data: List[Dict]) -> Dict[str, Any]:
        """Detect potential accuracy issues in the data."""
        issues = {
            'suspicious_patterns': 0,
            'potential_duplicates': 0,
            'outliers_detected': 0,
            'details': []
        }
        
        # Check for exact duplicate rows
        row_hashes = set()
        for row in preview_data:
            row_str = str(sorted(row.items()))
            if row_str in row_hashes:
                issues['potential_duplicates'] += 1
            else:
                row_hashes.add(row_str)
        
        # Check for suspicious patterns (all same values, sequential IDs, etc.)
        for column in preview_data[0].keys():
            values = [row[column] for row in preview_data if row[column] is not None]
            if values:
                unique_values = set(str(v) for v in values)
                
                # All same values
                if len(unique_values) == 1 and len(values) > 10:
                    issues['suspicious_patterns'] += 1
                    issues['details'].append(f'Column {column}: All values are identical')
                
                # Sequential numbers that might indicate test data
                if all(isinstance(v, (int, float)) for v in values[:10]):
                    numeric_values = [float(v) for v in values[:10]]
                    if all(numeric_values[i] + 1 == numeric_values[i+1] for i in range(len(numeric_values)-1)):
                        issues['suspicious_patterns'] += 1
                        issues['details'].append(f'Column {column}: Sequential numeric pattern detected')
        
        return issues
    
    def _generate_preview_insights(self, preview_data: List[Dict], config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate insights about the preview data for users."""
        insights = []
        
        if not preview_data:
            insights.append({
                'type': 'warning',
                'title': 'No Data Found',
                'message': 'The current configuration produces no results. Check your join conditions and filters.',
                'action': 'Review join configuration'
            })
            return insights
        
        # Data volume insights
        row_count = len(preview_data)
        if row_count == 100:  # Likely hit the limit
            insights.append({
                'type': 'info',
                'title': 'Preview Limited',
                'message': f'Showing first {row_count} rows. Full dataset may be larger.',
                'action': 'Consider adding filters to reduce result size'
            })
        
        # Column insights
        columns = list(preview_data[0].keys())
        if len(columns) > 15:
            insights.append({
                'type': 'warning',
                'title': 'Many Columns',
                'message': f'Result has {len(columns)} columns. Consider selecting specific columns for better performance.',
                'action': 'Use column selection to focus on needed data'
            })
        
        # Data quality insights
        null_percentage = sum(1 for row in preview_data for value in row.values() if value is None) / (len(preview_data) * len(columns)) * 100
        if null_percentage > 30:
            insights.append({
                'type': 'warning',
                'title': 'Missing Data',
                'message': f'{null_percentage:.1f}% of cells are empty. This may affect analysis.',
                'action': 'Consider data cleaning or filtering out incomplete records'
            })
        
        # Join effectiveness
        if len(config.get('tables', [])) > 1:
            estimated_reduction = self._estimate_join_reduction(config)
            if estimated_reduction > 0.5:
                insights.append({
                    'type': 'success',
                    'title': 'Effective Join',
                    'message': f'Join appears to be filtering data effectively (≈{estimated_reduction*100:.0f}% reduction).',
                    'action': 'Good filtering - data is well-matched'
                })
            elif estimated_reduction < 0.1:
                insights.append({
                    'type': 'warning',
                    'title': 'Minimal Join Effect',
                    'message': 'Join conditions may be too broad, producing many results.',
                    'action': 'Consider adding more restrictive join conditions'
                })
        
        return insights
    
    def _estimate_join_reduction(self, config: Dict[str, Any]) -> float:
        """Estimate how much the joins reduce the data size."""
        # Simple heuristic - would need actual table sizes for accurate calculation
        tables = config.get('tables', [])
        if len(tables) <= 1:
            return 0.0
        
        # Estimate based on join types
        reduction_factor = 0.3  # Default assumption
        for join in config.get('joins', []):
            join_type = join.get('join_type', 'INNER').upper()
            if join_type == 'INNER':
                reduction_factor += 0.2
            elif join_type in ['LEFT', 'RIGHT']:
                reduction_factor += 0.1
        
        return min(reduction_factor, 0.9)
    
    def _generate_export_recommendations(self, preview_data: List[Dict], config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate recommendations for exporting the data."""
        recommendations = []
        
        if not preview_data:
            return recommendations
        
        # Size recommendations
        estimated_size = len(preview_data) * 10  # Rough estimate
        if estimated_size > 1000:
            recommendations.append({
                'type': 'performance',
                'title': 'Large Dataset',
                'message': 'Consider exporting in chunks or adding filters to reduce size.',
                'formats': ['csv', 'json'],
                'priority': 'high'
            })
        
        # Format recommendations based on data characteristics
        columns = list(preview_data[0].keys())
        has_long_text = any(
            len(str(value)) > 100 
            for row in preview_data[:10] 
            for value in row.values() 
            if value is not None
        )
        
        if has_long_text:
            recommendations.append({
                'type': 'format',
                'title': 'Long Text Content',
                'message': 'Excel format recommended for long text content with proper cell wrapping.',
                'formats': ['excel'],
                'priority': 'medium'
            })
        
        if len(columns) <= 5 and len(preview_data) <= 100:
            recommendations.append({
                'type': 'format',
                'title': 'Small Clean Dataset',
                'message': 'Perfect for PDF export with formatting.',
                'formats': ['pdf', 'excel'],
                'priority': 'low'
            })
        
        # Special character recommendations
        has_special_chars = any(
            any(char in str(value) for char in ['\n', '\r', '\t', ',', '"']) 
            for row in preview_data[:10] 
            for value in row.values() 
            if value is not None
        )
        
        if has_special_chars:
            recommendations.append({
                'type': 'format',
                'title': 'Special Characters Detected',
                'message': 'Use Excel or JSON format to preserve special characters properly.',
                'formats': ['excel', 'json'],
                'priority': 'medium'
            })
        
        return recommendations
    
    def _get_tables_info(self, tables: List[str]) -> Dict[str, Any]:
        """Get detailed information about the tables involved."""
        info = {}
        
        for table in tables:
            try:
                # Get column info
                columns = self.get_table_columns(table)
                
                # Get row count
                count_result = db.session.execute(text(f"SELECT COUNT(*) FROM {table}"))
                row_count = count_result.scalar()
                
                # Get sample data
                sample_result = db.session.execute(text(f"SELECT * FROM {table} LIMIT 3"))
                sample_data = []
                for row in sample_result:
                    sample_data.append({columns[i]['name']: value for i, value in enumerate(row)})
                
                info[table] = {
                    'columns': columns,
                    'row_count': row_count,
                    'sample_data': sample_data
                }
                
            except Exception as e:
                info[table] = {'error': str(e)}
        
        return info
    
    def _analyze_joins(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze the joins and provide insights."""
        analysis = {
            'join_count': len(config.get('joins', [])),
            'join_details': [],
            'potential_issues': [],
            'recommendations': []
        }
        
        for join in config.get('joins', []):
            try:
                table1 = join['table1']
                table2 = join['table2']
                column1 = join['column1']
                column2 = join['column2']
                join_type = join.get('join_type', 'INNER')
                
                # Analyze join cardinality
                count1_query = f"SELECT COUNT(DISTINCT {column1}) FROM {table1}"
                count2_query = f"SELECT COUNT(DISTINCT {column2}) FROM {table2}"
                
                distinct1 = db.session.execute(text(count1_query)).scalar()
                distinct2 = db.session.execute(text(count2_query)).scalar()
                
                # Check for potential many-to-many relationships
                join_analysis = {
                    'table1': table1,
                    'table2': table2,
                    'column1': column1,
                    'column2': column2,
                    'join_type': join_type,
                    'distinct_values_1': distinct1,
                    'distinct_values_2': distinct2,
                    'estimated_cardinality': self._estimate_join_cardinality(table1, column1, table2, column2)
                }
                
                analysis['join_details'].append(join_analysis)
                
                # Add recommendations based on analysis
                if distinct1 > distinct2 * 10:
                    analysis['recommendations'].append(f"Consider indexing {table2}.{column2} for better performance")
                elif distinct2 > distinct1 * 10:
                    analysis['recommendations'].append(f"Consider indexing {table1}.{column1} for better performance")
                
            except Exception as e:
                analysis['potential_issues'].append(f"Error analyzing join {table1}.{column1} = {table2}.{column2}: {str(e)}")
        
        return analysis
    
    def _estimate_join_cardinality(self, table1: str, column1: str, table2: str, column2: str) -> str:
        """Estimate the cardinality of a join relationship."""
        try:
            # Check if it's one-to-one, one-to-many, or many-to-many
            query1 = f"SELECT COUNT(*) as total, COUNT(DISTINCT {column1}) as distinct_vals FROM {table1}"
            query2 = f"SELECT COUNT(*) as total, COUNT(DISTINCT {column2}) as distinct_vals FROM {table2}"
            
            result1 = db.session.execute(text(query1)).fetchone()
            result2 = db.session.execute(text(query2)).fetchone()
            
            total1, distinct1 = result1[0], result1[1]
            total2, distinct2 = result2[0], result2[1]
            
            # Simple heuristic for cardinality estimation
            if distinct1 == total1 and distinct2 == total2:
                return "one-to-one"
            elif distinct1 == total1:
                return "one-to-many"
            elif distinct2 == total2:
                return "many-to-one"
            else:
                return "many-to-many"
                
        except Exception:
            return "unknown"
    
    def _build_join_query(self, config: Dict[str, Any], limit: Optional[int] = None) -> str:
        """Build SQL query for joining tables based on configuration."""
        tables = config.get('tables', [])
        joins = config.get('joins', [])
        filters = config.get('filters', [])
        selected_columns = config.get('selected_columns', {})
        
        if not tables:
            raise ValueError("No tables specified")
        
        # Start with the first table
        main_table = tables[0]
        query_parts = []
        
        # Build SELECT clause
        select_columns = []
        for table in tables:
            table_columns = selected_columns.get(table, [])
            if not table_columns:
                # If no specific columns selected, select all
                table_columns = [col['name'] for col in self.get_table_columns(table)]
            
            for col in table_columns:
                # For single table, don't prefix with table name to avoid redundancy
                if len(tables) == 1:
                    select_columns.append(f"{col}")
                else:
                    # Avoid duplicate column names by prefixing with table name
                    select_columns.append(f"{table}.{col} AS {table}_{col}")
        
        query_parts.append(f"SELECT {', '.join(select_columns)}")
        query_parts.append(f"FROM {main_table}")
        
        # Add joins
        for join in joins:
            table1 = join['table1']
            table2 = join['table2']
            column1 = join['column1']
            column2 = join['column2']
            join_type = join.get('join_type', 'INNER')
            condition_type = join.get('condition_type', '=')
            
            join_condition = f"{table1}.{column1} {condition_type} {table2}.{column2}"
            query_parts.append(f"{join_type} JOIN {table2} ON {join_condition}")
        
        # Build WHERE clause from filters and custom where
        where_conditions = []
        
        # Add filter conditions
        for filter_rule in filters:
            condition = self._build_filter_condition(filter_rule)
            if condition:
                where_conditions.append(condition)
        
        # Add custom WHERE clause if specified
        custom_where = config.get('where_clause', '').strip()
        if custom_where:
            where_conditions.append(f"({custom_where})")
        
        # Add WHERE clause if we have conditions
        if where_conditions:
            query_parts.append(f"WHERE {' AND '.join(where_conditions)}")
        
        # Add ORDER BY if specified
        order_by = config.get('order_by', '').strip()
        if order_by:
            query_parts.append(f"ORDER BY {order_by}")
        
        # Add LIMIT if specified
        if limit:
            query_parts.append(f"LIMIT {limit}")
        
        return ' '.join(query_parts)
    
    def _build_filter_condition(self, filter_rule: Dict[str, Any]) -> str:
        """Build a SQL condition from a filter rule."""
        table = filter_rule.get('table')
        column = filter_rule.get('column')
        operator = filter_rule.get('operator')
        value = filter_rule.get('value', '')
        value2 = filter_rule.get('value2', '')
        data_type = filter_rule.get('data_type', 'text')
        
        if not table or not column or not operator:
            return ''
        
        column_ref = f"{table}.{column}"
        
        # Handle NULL checks
        if operator == 'IS NULL':
            return f"{column_ref} IS NULL"
        elif operator == 'IS NOT NULL':
            return f"{column_ref} IS NOT NULL"
        
        if not value:
            return ''
        
        # Format value based on data type
        if data_type == 'text':
            escaped_value = value.replace("'", "''")  # Escape single quotes
            if operator == 'LIKE':
                return f"{column_ref} LIKE '%{escaped_value}%'"
            elif operator == 'NOT LIKE':
                return f"{column_ref} NOT LIKE '%{escaped_value}%'"
            elif operator == 'STARTS_WITH':
                return f"{column_ref} LIKE '{escaped_value}%'"
            elif operator == 'ENDS_WITH':
                return f"{column_ref} LIKE '%{escaped_value}'"
            else:
                return f"{column_ref} {operator} '{escaped_value}'"
        
        elif data_type == 'number':
            try:
                num_value = float(value)
                if operator == 'BETWEEN':
                    if value2:
                        num_value2 = float(value2)
                        return f"{column_ref} BETWEEN {num_value} AND {num_value2}"
                elif operator == 'NOT BETWEEN':
                    if value2:
                        num_value2 = float(value2)
                        return f"{column_ref} NOT BETWEEN {num_value} AND {num_value2}"
                else:
                    return f"{column_ref} {operator} {num_value}"
            except ValueError:
                return ''
        
        elif data_type == 'date':
            # Validate date format (basic validation)
            if len(value) == 10 and value.count('-') == 2:
                if operator == 'BETWEEN':
                    if value2 and len(value2) == 10 and value2.count('-') == 2:
                        return f"{column_ref} BETWEEN '{value}' AND '{value2}'"
                elif operator == 'NOT BETWEEN':
                    if value2 and len(value2) == 10 and value2.count('-') == 2:
                        return f"{column_ref} NOT BETWEEN '{value}' AND '{value2}'"
                else:
                    return f"{column_ref} {operator} '{value}'"
        
        return ''
    
    def export_joined_data_to_excel(self, config: Dict[str, Any], filename: Optional[str] = None) -> Dict[str, Any]:
        """Export joined data to Excel file."""
        try:
            # Get the full dataset (no limit)
            query = self._build_join_query(config)
            result = db.session.execute(text(query))
            columns = list(result.keys())
            rows = result.fetchall()
            
            # Convert to pandas DataFrame
            data = []
            for row in rows:
                row_dict = {}
                for i, value in enumerate(row):
                    row_dict[columns[i]] = value
                data.append(row_dict)
            
            df = pd.DataFrame(data)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                table_names = "_".join(config.get('tables', []))[:50]  # Limit length
                filename = f"joined_data_{table_names}_{timestamp}.xlsx"
            
            # Create Excel file in uploads directory
            uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
            os.makedirs(uploads_dir, exist_ok=True)
            file_path = os.path.join(uploads_dir, filename)
            
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # Main data sheet
            ws_data = wb.active
            ws_data.title = "Joined Data"
            
            # Add headers with styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(columns, 1):
                cell = ws_data.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add data
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(columns, 1):
                    ws_data.cell(row=row_num, column=col_num, value=row_data.get(header))
            
            # Auto-adjust column widths
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Add configuration sheet
            ws_config = wb.create_sheet("Configuration")
            config_data = [
                ["Join Configuration"],
                [""],
                ["Tables:", ", ".join(config.get('tables', []))],
                ["Join Type:", config.get('joins', [{}])[0].get('join_type', 'N/A') if config.get('joins') else 'N/A'],
                ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Records:", len(data)],
                [""],
                ["Joins:"]
            ]
            
            for join in config.get('joins', []):
                config_data.append([
                    f"{join['table1']}.{join['column1']} = {join['table2']}.{join['column2']} ({join.get('join_type', 'INNER')})"
                ])
            
            for row_num, row_data in enumerate(config_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    ws_config.cell(row=row_num, column=col_num, value=value)
            
            # Save the workbook
            wb.save(file_path)
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'record_count': len(data),
                'column_count': len(columns)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_table_sample_data(self, table_name: str, limit: int = 5) -> Dict[str, Any]:
        """Get sample data from a table for relationship analysis."""
        try:
            inspector = inspect(db.engine)
            if table_name not in inspector.get_table_names():
                return {'success': False, 'error': 'Table not found'}
            
            query = f"SELECT * FROM {table_name} LIMIT {limit}"
            result = db.session.execute(text(query))
            columns = list(result.keys())
            rows = result.fetchall()
            
            data = []
            for row in rows:
                row_dict = {}
                for i, value in enumerate(row):
                    row_dict[columns[i]] = value
                data.append(row_dict)
            
            return {
                'success': True,
                'columns': columns,
                'data': data
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def execute_custom_query(self, query: str, database: str = 'excel_data') -> Dict[str, Any]:
        """Execute a custom SQL query safely."""
        try:
            # Basic security check - only allow SELECT statements
            query_upper = query.strip().upper()
            if not query_upper.startswith('SELECT'):
                raise ValueError("Only SELECT queries are allowed")
            
            # Additional security checks
            dangerous_keywords = ['DROP', 'DELETE', 'UPDATE', 'INSERT', 'ALTER', 'CREATE', 'EXEC', 'EXECUTE']
            for keyword in dangerous_keywords:
                if keyword in query_upper:
                    raise ValueError(f"Query contains forbidden keyword: {keyword}")
            
            # Get database path
            if database == 'stock':
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'Stock.db')
            else:
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', f'{database}.db')
            
            if not os.path.exists(db_path):
                raise Exception(f"Database {database} not found")
            
            # Execute the query using direct SQLite connection
            import sqlite3
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row  # This allows us to access columns by name
            cursor = conn.cursor()
            
            # Execute the query
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # Get column names
            columns = [description[0] for description in cursor.description] if cursor.description else []
            
            # Convert rows to list of dictionaries
            data = [dict(row) for row in rows]
            
            conn.close()
            
            return {
                'data': data,
                'columns': columns,
                'row_count': len(data)
            }
            
        except Exception as e:
            raise Exception(f"Error executing custom query: {str(e)}")
    
    def save_relationship_configuration(self, config: Dict[str, Any], name: str, database: str = 'stock') -> Dict[str, Any]:
        """Save a relationship configuration for reuse."""
        try:
            # Get database path
            if database == 'stock':
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'Stock.db')
            else:
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', f'{database}.db')
            
            if not os.path.exists(db_path):
                raise Exception(f"Database {database} not found")
            
            import sqlite3
            import json
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Create configurations table if it doesn't exist
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS relationship_configurations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    configuration TEXT NOT NULL,
                    created_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_date DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            config_json = json.dumps(config)
            
            # Check if configuration name already exists
            cursor.execute("SELECT id FROM relationship_configurations WHERE name = ?", (name,))
            existing = cursor.fetchone()
            
            if existing:
                # Update existing configuration
                cursor.execute('''
                    UPDATE relationship_configurations 
                    SET configuration = ?, updated_date = CURRENT_TIMESTAMP 
                    WHERE name = ?
                ''', (config_json, name))
                message = f"Configuration '{name}' updated successfully"
            else:
                # Insert new configuration
                cursor.execute('''
                    INSERT INTO relationship_configurations (name, configuration) 
                    VALUES (?, ?)
                ''', (name, config_json))
                message = f"Configuration '{name}' saved successfully"
            
            conn.commit()
            conn.close()
            
            return {
                'success': True,
                'message': message
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': str(e)
            }
    
    def load_relationship_configuration(self, name: str) -> Dict[str, Any]:
        """Load a saved relationship configuration."""
        try:
            result = db.session.execute(text(
                "SELECT configuration FROM relationship_configurations WHERE name = :name"
            ), {'name': name}).fetchone()
            
            if not result:
                return {
                    'success': False,
                    'error': f"Configuration '{name}' not found"
                }
            
            import json
            config = json.loads(result[0])
            
            return {
                'success': True,
                'configuration': config
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def list_saved_configurations(self) -> Dict[str, Any]:
        """List all saved relationship configurations."""
        try:
            results = db.session.execute(text('''
                SELECT name, created_date, updated_date 
                FROM relationship_configurations 
                ORDER BY updated_date DESC
            ''')).fetchall()
            
            configurations = []
            for row in results:
                configurations.append({
                    'name': row[0],
                    'created_date': row[1],
                    'updated_date': row[2]
                })
            
            return {
                'success': True,
                'configurations': configurations
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'configurations': []
            }
    
    def delete_relationship_configuration(self, name: str) -> Dict[str, Any]:
        """Delete a saved relationship configuration."""
        try:
            result = db.session.execute(text(
                "DELETE FROM relationship_configurations WHERE name = :name"
            ), {'name': name})
            
            if result.rowcount == 0:
                return {
                    'success': False,
                    'error': f"Configuration '{name}' not found"
                }
            
            db.session.commit()
            
            return {
                'success': True,
                'message': f"Configuration '{name}' deleted successfully"
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': str(e)
            }
    
    def test_relationship_performance(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Test the performance of a relationship configuration."""
        try:
            import time
            
            # Test query execution time
            start_time = time.time()
            query = self._build_join_query(config, limit=1000)
            result = db.session.execute(text(query))
            rows = result.fetchall()
            execution_time = time.time() - start_time
            
            # Analyze query complexity
            complexity_analysis = {
                'tables_count': len(config.get('tables', [])),
                'joins_count': len(config.get('joins', [])),
                'filters_count': len(config.get('filters', [])),
                'estimated_complexity': 'low'
            }
            
            # Simple complexity estimation
            total_complexity = (
                complexity_analysis['tables_count'] * 2 +
                complexity_analysis['joins_count'] * 3 +
                complexity_analysis['filters_count'] * 1
            )
            
            if total_complexity > 15:
                complexity_analysis['estimated_complexity'] = 'high'
            elif total_complexity > 8:
                complexity_analysis['estimated_complexity'] = 'medium'
            
            return {
                'success': True,
                'execution_time': round(execution_time, 3),
                'row_count': len(rows),
                'complexity_analysis': complexity_analysis,
                'query': query,
                'recommendations': self._get_performance_recommendations(complexity_analysis, execution_time)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _get_performance_recommendations(self, complexity: Dict[str, Any], execution_time: float) -> List[str]:
        """Get performance recommendations based on analysis."""
        recommendations = []
        
        if execution_time > 2.0:
            recommendations.append("Query execution time is slow. Consider adding indexes on join columns.")
        
        if complexity['joins_count'] > 3:
            recommendations.append("Multiple joins detected. Consider breaking down into smaller queries if performance is poor.")
        
        if complexity['estimated_complexity'] == 'high':
            recommendations.append("High complexity query. Monitor performance and consider optimization.")
        
        if not recommendations:
            recommendations.append("Query performance looks good!")
        
        return recommendations
    
    def execute_query_as_table(self, query: str, table_name: str, database: str = 'excel_data') -> Dict[str, Any]:
        """Execute a SQL query and save the results as a new table."""
        try:
            # Get database path
            if database == 'stock':
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'Stock.db')
            else:
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', f'{database}.db')
            
            if not os.path.exists(db_path):
                return {
                    'success': False,
                    'error': f'Database {database} not found'
                }
            
            import sqlite3
            import pandas as pd
            
            # First, execute the query to get results
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute(query)
            rows = cursor.fetchall()
            columns = [description[0] for description in cursor.description] if cursor.description else []
            
            if not rows:
                conn.close()
                return {
                    'success': False,
                    'error': 'Query returned no results'
                }
            
            # Check if table already exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
            if cursor.fetchone():
                conn.close()
                return {
                    'success': False,
                    'error': f'Table "{table_name}" already exists. Please choose a different name.'
                }
            
            # Convert results to list of dictionaries
            data = [dict(row) for row in rows]
            
            # Convert to pandas DataFrame for easier handling
            df = pd.DataFrame(data)
            
            # Create table from DataFrame using pandas to_sql
            df.to_sql(table_name, conn, index=False, if_exists='fail')
            
            conn.commit()
            
            # Get the newly created table info
            new_table_columns = self.get_table_columns(table_name, database)
            
            conn.close()
            
            return {
                'success': True,
                'message': f'Table "{table_name}" created successfully with {len(rows)} rows',
                'table_name': table_name,
                'row_count': len(rows),
                'columns': new_table_columns,
                'sample_data': data[:5]  # First 5 rows as sample
            }
                
        except Exception as e:
            return {
                'success': False,
                'error': f'Error creating table: {str(e)}'
            }
    
    def analyze_table_data(self, table_name: str, database: str = 'excel_data') -> Dict[str, Any]:
        """Perform comprehensive analysis of table data."""
        try:
            # Handle different database connections
            if database and database != 'excel_data':
                # Connect to specific database
                import sqlite3
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', f'{database}.db')
                
                if not os.path.exists(db_path):
                    return {
                        'success': False,
                        'error': f'Database "{database}" not found'
                    }
                
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row  # This allows accessing columns by name
                
                # Check if table exists
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                if not cursor.fetchone():
                    conn.close()
                    return {
                        'success': False,
                        'error': f'Table "{table_name}" not found in database "{database}"'
                    }
                
                # Get table data
                df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
                
                # Get column info from sqlite
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()
                columns = [
                    {
                        'name': col[1],
                        'type': col[2],
                        'nullable': not col[3],
                        'primary_key': col[5] == 1
                    }
                    for col in columns_info
                ]
                
                conn.close()
            else:
                # Use default database connection
                inspector = inspect(db.engine)
                if table_name not in inspector.get_table_names():
                    return {
                        'success': False,
                        'error': f'Table "{table_name}" not found'
                    }
                
                with db.engine.connect() as connection:
                    # Get basic table info
                    result = connection.execute(text(f"SELECT * FROM {table_name}"))
                    rows = result.fetchall()
                    column_names = list(result.keys())
                    
                    if not rows:
                        return {
                            'success': False,
                            'error': 'Table is empty'
                        }
                    
                    # Convert to DataFrame for analysis
                    df = pd.DataFrame(rows, columns=column_names)
                    columns = self.get_table_columns(table_name)
            
            if df.empty:
                return {
                    'success': False,
                    'error': 'Table is empty'
                }
            
            # Perform comprehensive analysis
            analysis = {
                'success': True,
                'table_name': table_name,
                'database': database,
                'row_count': len(df),
                'columns': columns,
                'column_types': self._analyze_column_types(df),
                'sample_data': self._convert_numpy_types(df.head(10).to_dict('records')),
                'missing_analysis': self._analyze_missing_data(df),
                'duplicate_analysis': self._analyze_duplicates(df),
                'numeric_stats': self._analyze_numeric_stats(df),
                'text_stats': self._analyze_text_stats(df),
                'date_stats': self._analyze_date_stats(df),
                'value_frequencies': self._analyze_value_frequencies(df),
                'unique_analysis': self._analyze_unique_values(df),
                'numeric_distributions': self._analyze_numeric_distributions(df),
                'data_issues': self._detect_data_issues(df),
                'quality_score': self._calculate_quality_score(df),
                'text_patterns': self._analyze_text_patterns(df),
                'trends': self._analyze_trends(df),
                'anomalies': self._detect_anomalies(df),
                'correlations': self._analyze_correlations(df),
                'relationship_suggestions': self._suggest_relationships(df),
                'correlation_matrix': self._create_correlation_matrix(df)
            }
            
            return self._convert_numpy_types(analysis)
                
        except Exception as e:
            return {
                'success': False,
                'error': f'Error analyzing table: {str(e)}'
            }
    
    def _analyze_column_types(self, df: pd.DataFrame) -> Dict[str, int]:
        """Analyze column data types distribution."""
        type_counts = {}
        for column in df.columns:
            dtype_str = str(df[column].dtype)
            if dtype_str.startswith('int') or dtype_str.startswith('float'):
                type_name = 'Numeric'
            elif dtype_str.startswith('object'):
                type_name = 'Text'
            elif dtype_str.startswith('datetime'):
                type_name = 'Date/Time'
            elif dtype_str.startswith('bool'):
                type_name = 'Boolean'
            else:
                type_name = 'Other'
            
            type_counts[type_name] = type_counts.get(type_name, 0) + 1
        
        return type_counts
    
    def _analyze_missing_data(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze missing data patterns."""
        missing_analysis = {}
        total_rows = len(df)
        
        for column in df.columns:
            null_count = df[column].isnull().sum()
            missing_analysis[column] = {
                'count': int(null_count),
                'percentage': (null_count / total_rows) * 100 if total_rows > 0 else 0
            }
        
        return missing_analysis
    
    def _analyze_duplicates(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze duplicate data."""
        duplicate_rows = df.duplicated().sum()
        total_rows = len(df)
        
        return {
            'duplicate_rows': int(duplicate_rows),
            'duplicate_percentage': (duplicate_rows / total_rows) * 100 if total_rows > 0 else 0,
            'unique_rows': total_rows - duplicate_rows
        }
    
    def _analyze_numeric_stats(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze numeric column statistics."""
        numeric_stats = {}
        numeric_columns = df.select_dtypes(include=['number']).columns
        
        for column in numeric_columns:
            series = df[column].dropna()
            if len(series) > 0:
                numeric_stats[column] = {
                    'mean': float(series.mean()),
                    'median': float(series.median()),
                    'std': float(series.std()) if len(series) > 1 else 0,
                    'min': float(series.min()),
                    'max': float(series.max()),
                    'range': float(series.max() - series.min()),
                    'q25': float(series.quantile(0.25)),
                    'q75': float(series.quantile(0.75))
                }
        
        return numeric_stats
    
    def _analyze_text_stats(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze text column statistics."""
        text_stats = {}
        text_columns = df.select_dtypes(include=['object']).columns
        
        for column in text_columns:
            series = df[column].dropna().astype(str)
            if len(series) > 0:
                lengths = series.str.len()
                text_stats[column] = {
                    'avg_length': float(lengths.mean()),
                    'min_length': int(lengths.min()),
                    'max_length': int(lengths.max()),
                    'unique_count': int(series.nunique()),
                    'empty_count': int((series == '').sum())
                }
        
        return text_stats
    
    def _analyze_date_stats(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze date/time column statistics."""
        date_stats = {}
        
        # Try to identify date columns
        for column in df.columns:
            try:
                # Try to convert to datetime
                date_series = pd.to_datetime(df[column], errors='coerce').dropna()
                if len(date_series) > len(df) * 0.5:  # If more than 50% are valid dates
                    date_stats[column] = {
                        'min_date': str(date_series.min().date()),
                        'max_date': str(date_series.max().date()),
                        'date_range': str((date_series.max() - date_series.min()).days) + ' days'
                    }
            except:
                continue
        
        return date_stats
    
    def _analyze_value_frequencies(self, df: pd.DataFrame) -> Dict[str, Dict[str, int]]:
        """Analyze value frequencies for each column."""
        frequencies = {}
        
        for column in df.columns:
            # Limit to top 10 most frequent values
            value_counts = df[column].value_counts().head(10)
            frequencies[column] = {str(k): int(v) for k, v in value_counts.items()}
        
        return frequencies
    
    def _analyze_unique_values(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze unique value patterns."""
        unique_analysis = {}
        total_rows = len(df)
        
        for column in df.columns:
            unique_count = df[column].nunique()
            unique_analysis[column] = {
                'unique_count': int(unique_count),
                'uniqueness_ratio': unique_count / total_rows if total_rows > 0 else 0
            }
        
        return unique_analysis
    
    def _analyze_numeric_distributions(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Analyze numeric distributions."""
        distributions = {}
        numeric_columns = df.select_dtypes(include=['number']).columns
        
        for column in numeric_columns:
            series = df[column].dropna()
            if len(series) > 1:
                try:
                    from scipy import stats
                    distributions[column] = {
                        'skewness': float(stats.skew(series)),
                        'kurtosis': float(stats.kurtosis(series)),
                        'zero_count': int((series == 0).sum())
                    }
                except ImportError:
                    # Fallback without scipy
                    distributions[column] = {
                        'skewness': 'N/A (scipy not available)',
                        'kurtosis': 'N/A (scipy not available)',
                        'zero_count': int((series == 0).sum())
                    }
        
        return distributions
    
    def _detect_data_issues(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Detect various data quality issues."""
        issues = []
        
        # Check for columns with high missing data
        for column in df.columns:
            missing_pct = (df[column].isnull().sum() / len(df)) * 100
            if missing_pct > 50:
                issues.append({
                    'type': 'High Missing Data',
                    'severity': 'high',
                    'description': f'Column "{column}" has {missing_pct:.1f}% missing values',
                    'count': int(df[column].isnull().sum())
                })
        
        # Check for duplicate rows
        duplicate_count = df.duplicated().sum()
        if duplicate_count > 0:
            issues.append({
                'type': 'Duplicate Rows',
                'severity': 'medium',
                'description': f'Found {duplicate_count} duplicate rows',
                'count': int(duplicate_count)
            })
        
        # Check for columns with single unique value
        for column in df.columns:
            if df[column].nunique() == 1:
                issues.append({
                    'type': 'Constant Column',
                    'severity': 'low',
                    'description': f'Column "{column}" has only one unique value',
                    'count': 1
                })
        
        return issues
    
    def _calculate_quality_score(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate overall data quality score."""
        total_rows = len(df)
        total_cells = total_rows * len(df.columns)
        
        # Calculate various quality metrics
        missing_cells = df.isnull().sum().sum()
        duplicate_rows = df.duplicated().sum()
        
        # Completeness score (100% - missing percentage)
        completeness = ((total_cells - missing_cells) / total_cells) * 100 if total_cells > 0 else 0
        
        # Uniqueness score (100% - duplicate percentage)
        uniqueness = ((total_rows - duplicate_rows) / total_rows) * 100 if total_rows > 0 else 0
        
        # Consistency score (simplified - based on data type consistency)
        consistency = 85  # Placeholder - would need more complex analysis
        
        # Overall score (weighted average)
        overall_score = (completeness * 0.4 + uniqueness * 0.3 + consistency * 0.3)
        
        return {
            'overall_score': overall_score,
            'breakdown': {
                'completeness': completeness,
                'uniqueness': uniqueness,
                'consistency': consistency
            }
        }
    
    def _analyze_text_patterns(self, df: pd.DataFrame) -> Dict[str, List[Dict[str, Any]]]:
        """Analyze text patterns in string columns."""
        patterns = {}
        text_columns = df.select_dtypes(include=['object']).columns
        
        for column in text_columns:
            series = df[column].dropna().astype(str)
            if len(series) > 0:
                # Simple pattern analysis
                pattern_counts = {}
                for value in series.head(100):  # Limit to first 100 values
                    # Extract basic patterns
                    if value.isdigit():
                        pattern = 'All digits'
                    elif value.isalpha():
                        pattern = 'All letters'
                    elif '@' in value:
                        pattern = 'Email-like'
                    elif len(value.split()) > 1:
                        pattern = 'Multiple words'
                    else:
                        pattern = 'Mixed/Other'
                    
                    pattern_counts[pattern] = pattern_counts.get(pattern, 0) + 1
                
                patterns[column] = [
                    {'pattern': k, 'count': v} 
                    for k, v in sorted(pattern_counts.items(), key=lambda x: x[1], reverse=True)[:5]
                ]
        
        return patterns
    
    def _analyze_trends(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Analyze data trends."""
        trends = []
        
        # Simple trend analysis for numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns
        for column in numeric_columns:
            series = df[column].dropna()
            if len(series) > 1:
                # Check if data is increasing or decreasing
                diff = series.diff().dropna()
                if len(diff) > 0:
                    increasing = (diff > 0).sum()
                    decreasing = (diff < 0).sum()
                    
                    if increasing > decreasing * 1.5:
                        trends.append({
                            'column': column,
                            'description': 'Generally increasing trend',
                            'confidence': int((increasing / len(diff)) * 100)
                        })
                    elif decreasing > increasing * 1.5:
                        trends.append({
                            'column': column,
                            'description': 'Generally decreasing trend',
                            'confidence': int((decreasing / len(diff)) * 100)
                        })
        
        return trends
    
    def _detect_anomalies(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Detect anomalies in the data."""
        anomalies = []
        
        # Detect outliers in numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns
        for column in numeric_columns:
            series = df[column].dropna()
            if len(series) > 4:
                Q1 = series.quantile(0.25)
                Q3 = series.quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                
                outliers = series[(series < lower_bound) | (series > upper_bound)]
                if len(outliers) > 0:
                    anomalies.append({
                        'type': 'Statistical Outliers',
                        'column': column,
                        'description': f'Found {len(outliers)} outliers using IQR method',
                        'count': len(outliers)
                    })
        
        return anomalies
    
    def _analyze_correlations(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Analyze correlations between numeric columns."""
        correlations = []
        numeric_df = df.select_dtypes(include=['number'])
        
        if len(numeric_df.columns) > 1:
            corr_matrix = numeric_df.corr()
            
            # Extract significant correlations
            for i, col1 in enumerate(corr_matrix.columns):
                for j, col2 in enumerate(corr_matrix.columns):
                    if i < j:  # Avoid duplicates
                        corr_value = corr_matrix.loc[col1, col2]
                        if not pd.isna(corr_value) and abs(corr_value) > 0.3:
                            correlations.append({
                                'column1': col1,
                                'column2': col2,
                                'value': float(corr_value)
                            })
        
        return correlations
    
    def _suggest_relationships(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Suggest potential relationships in the data."""
        suggestions = []
        
        # Look for potential foreign key relationships
        for column in df.columns:
            unique_ratio = df[column].nunique() / len(df)
            
            if 0.1 < unique_ratio < 0.9:  # Not too unique, not too repetitive
                suggestions.append({
                    'type': 'Potential Lookup Column',
                    'description': f'Column "{column}" might be a foreign key or categorical variable',
                    'confidence': int((1 - abs(unique_ratio - 0.5)) * 100)
                })
        
        return suggestions
    
    def _create_correlation_matrix(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Create a correlation matrix for numeric columns."""
        numeric_df = df.select_dtypes(include=['number'])
        
        if len(numeric_df.columns) > 1:
            corr_matrix = numeric_df.corr()
            return {
                'columns': list(corr_matrix.columns),
                'data': corr_matrix.values.tolist()
            }
        
        return None
